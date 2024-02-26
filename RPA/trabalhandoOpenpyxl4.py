from openpyxl import load_workbook
import os

arquivo = 'E:\\Faculdade\\Curso Python\\excell\\Vendedores.xlsx'
planilha = load_workbook(arquivo)

sheet_selecionada = planilha['Vendas']

somarAmandaMartins =0
somarElianeMoreira =0
somarLeonardoAlmeida =0
somarNicolasPereira =0

for linha in range(2, len(sheet_selecionada['A']) + 1):
    if sheet_selecionada['A%s' % linha].value == 'Amanda Martins':
        somarAmandaMartins = somarAmandaMartins + sheet_selecionada['C%s' % linha].value
    elif sheet_selecionada['A%s' % linha].value == 'Eliane Moreira':
        somarElianeMoreira = somarElianeMoreira + sheet_selecionada['C%s' % linha].value
    elif sheet_selecionada['A%s' % linha].value == 'Leonardo Almeida':
        somarLeonardoAlmeida = somarLeonardoAlmeida + sheet_selecionada['C%s' % linha].value
    elif sheet_selecionada['A%s' % linha].value == 'Nicolas Pereira':
        somarNicolasPereira = somarNicolasPereira + sheet_selecionada['C%s' % linha].value

sheet_resumo = planilha.create_sheet('Resumo')

sheet_resumo['A1'] = 'Vendedor'
sheet_resumo['B1'] = 'Vendas'

sheet_resumo['A2'] = 'Amanda Martins'
sheet_resumo['B2'] = somarAmandaMartins

sheet_resumo['A3'] = 'Eliane Moreira'
sheet_resumo['B3'] = somarElianeMoreira

sheet_resumo['A4'] = 'Leonardo Almeida'
sheet_resumo['B4'] = somarLeonardoAlmeida

sheet_resumo['A5'] = 'Nicolas Pereira'
sheet_resumo['B5'] = somarNicolasPereira


planilha.save(arquivo)

os.startfile(arquivo)