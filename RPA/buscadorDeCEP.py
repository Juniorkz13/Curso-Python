from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import pyautogui

webdriver = webdriver.Chrome()
webdriver.get('https://buscacepinter.correios.com.br/app/endereco/index.php')
pyautogui.sleep(3)
webdriver.find_element(By.ID, 'endereco').send_keys('31035430')
pyautogui.sleep(3)
webdriver.find_element(By.ID, 'btn_pesquisar').click()
pyautogui.sleep(3)
rua = webdriver.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[1]')[0].text
bairro = webdriver.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[2]')[0].text
cidade = webdriver.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[3]')[0].text
cep = webdriver.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[4]')[0].text
print(f'Rua: {rua} Bairro: {bairro} Cidade: {cidade} CEP: {cep}')

wb = Workbook()
ws = wb.active

ws['A1'] = "Rua"
ws['A2'] = rua
ws['B1'] = "Bairro"
ws['B2'] = bairro
ws['C1'] = "Cidade"
ws['C2'] = cidade
ws['D1'] = "CEP"
ws['D2'] = cep

wb.save("endereco.xlsx")
webdriver.quit()



