from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

arquivo = 'E:\Faculdade\Curso Python\excell\inserirDados.xlsx'
planilha = load_workbook(arquivo)

sheet_selecionada = planilha['Professor']

dadosTabela = [
    ['Nome', 'Idade', 'Salário'],
    ['João', 30, 1000],
    ['Maria', 25, 1200],
    ['José', 40, 1500],
    ['Ana', 35, 2000],
]

for linha in dadosTabela:
    sheet_selecionada.append(linha)

corTitulo = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

for celula in sheet_selecionada[1]:
    celula.fill = corTitulo

corCelulas = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')

for linha in sheet_selecionada.iter_rows(min_row=2, max_row=5, min_col=1, max_col=3):
    for celula in linha:
        celula.fill = corCelulas

planilha.save(arquivo)

os.startfile(arquivo)