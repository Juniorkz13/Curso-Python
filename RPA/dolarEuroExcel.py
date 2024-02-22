from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import Workbook


driver = webdriver.Chrome()

driver.get("https://www.google.com")

search_box = driver.find_element(By.NAME, "q")

search_box.send_keys("valor do dolar hoje")

search_box.send_keys(Keys.RETURN)

driver.implicitly_wait(10)

resultDolar = driver.find_element(By.XPATH, '//*[@id="knowledge-currency__updatable-data-column"]/div[1]/div[2]/span[1]').text

driver.get("https://www.google.com")

search_box = driver.find_element(By.NAME, "q")

search_box.send_keys("valor do euro hoje")

search_box.send_keys(Keys.RETURN)

driver.implicitly_wait(10)

resultEuro = driver.find_element(By.XPATH, '//*[@id="knowledge-currency__updatable-data-column"]/div[1]/div[2]/span[1]').text

driver.quit()

wb = Workbook()

ws = wb.active

ws['A1'] = "Valor do Dolar Hoje"
ws['A2'] = resultDolar
ws['B1'] = "Valor do Euro Hoje"
ws['B2'] = resultEuro


wb.save("valorDoDolarEEuro.xlsx")