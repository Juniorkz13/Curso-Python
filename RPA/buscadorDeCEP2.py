from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import Workbook
import pyautogui
import pandas as pd

webdriver = webdriver.Chrome()

df = pd.read_excel('E:\Faculdade\Curso Python\CEP.xlsx')

ceps_to_search = df['CEP'].tolist()


wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1).value = "Rua"
ws.cell(row=1, column=2).value = "Bairro"
ws.cell(row=1, column=3).value = "Cidade"
ws.cell(row=1, column=4).value = "CEP"

for cep in ceps_to_search:
    webdriver.get('https://buscacepinter.correios.com.br/app/endereco/index.php')
    pyautogui.sleep(3)
    webdriver.find_element(By.ID, 'endereco').send_keys(cep)
    pyautogui.sleep(3)
    webdriver.find_element(By.ID, 'btn_pesquisar').click()
    pyautogui.sleep(3)

    results = webdriver.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr')

    for result in results[:4]:
        rua = webdriver.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[1]')[0].text
        bairro = webdriver.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[2]')[0].text
        cidade = webdriver.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[3]')[0].text
        cep = webdriver.find_elements(By.XPATH, '//*[@id="resultado-DNEC"]/tbody/tr/td[4]')[0].text
        print(f'Rua: {rua} Bairro: {bairro} Cidade: {cidade} CEP: {cep}')
        webdriver.find_element(By.ID, 'btn_nbusca').click()


        ws.append([rua, bairro, cidade, cep])

wb.save("enderecosPorCEP2.xlsx")

webdriver.quit()



