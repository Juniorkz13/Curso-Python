from openpyxl import load_workbook
from openpyxl import Workbook
import os

arquivo = 'E:\\Faculdade\\Curso Python\\excell\\DadosSistema.xlsx'
planilha = load_workbook(arquivo)

sheet_selecionada = planilha['Dados']

novoArquivo = Workbook()
novaPlanilha = novoArquivo.active

for linha in sheet_selecionada:
    for celula in linha:
        novaPlanilha[celula.coordinate].value = celula.value

novaPlanilha.delete_rows(2)
novaPlanilha.delete_cols(2, 2)

novoArquivo.create_sheet('Resumo')

selecionaSheetResumo = novoArquivo['Resumo']

selecionaSheetResumo['A1'] = 'Vendedor'
selecionaSheetResumo['A2'] = 'Total Vendas'

selecionaSheetResumo['A2'] = 'Amanda Martins'
selecionaSheetResumo['B2'] = '=SUMIF(Dados!A:C,A2,Dados!C:C)'

selecionaSheetResumo['A3'] = 'Eliane Moreira'
selecionaSheetResumo['B3'] = '=SUMIF(Dados!A:C,A3,Dados!C:C)'

selecionaSheetResumo['A4'] = 'Leonardo Almeida'
selecionaSheetResumo['B4'] = '=SUMIF(Dados!A:C,A4,Dados!C:C)'

selecionaSheetResumo['A5'] = 'Nicolas Pereira'
selecionaSheetResumo['B5'] = '=SUMIF(Dados!A:C,A5,Dados!C:C)'

caminho = 'E:\\Faculdade\\Curso Python\\excell\\Resumo.xlsx'
novoArquivo.save(caminho)


os.startfile(caminho)